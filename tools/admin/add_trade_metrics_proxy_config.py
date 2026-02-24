"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds two proxy derived DATASETS rows for any instrument (e.g. ES, NQ) to
  run_config.xlsx:
    {ID}_FOOTPRINT_PROXY_1M  (dataset_type=derived_trade_metrics_proxy,
                              metric_type=footprint, source=1s OHLCV)
    {ID}_CVD_PROXY_1M        (dataset_type=derived_trade_metrics_proxy,
                              metric_type=cvd, source=1s OHLCV)
- Updates the matching INSTRUMENTS row to populate:
    footprint_proxy_dataset_id  -> {ID}_FOOTPRINT_PROXY_1M
    cvd_proxy_dataset_id        -> {ID}_CVD_PROXY_1M
- Re-exports the config snapshot.

The proxy tables are built from 1s OHLCV canonical data using Bulk Volume
Classification (BVC). They cover the full price history (e.g. 2016-present)
and complement the real footprint/CVD (2025-present) that are built from
trade-level data.

Instrument-agnostic: works for ES, NQ, or any future instrument.

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  # Add ES (uses default source DB_ES_OHLCV_1S):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_proxy_config.py --instrument-id ES

  # Add NQ (uses default source DB_NQ_OHLCV_1S):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_proxy_config.py --instrument-id NQ

  # Add with explicit source dataset id:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_proxy_config.py --instrument-id ES --source-dataset-id DB_ES_OHLCV_1S

What success looks like:
- Prints "Added {ID}_FOOTPRINT_PROXY_1M to DATASETS" (or "already exists, skipping").
- Prints "Added {ID}_CVD_PROXY_1M to DATASETS" (or "already exists, skipping").
- Prints "Updated INSTRUMENTS {ID} row: footprint_proxy_dataset_id, cvd_proxy_dataset_id."
- Prints "Config snapshot re-exported."

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
- "INSTRUMENTS row not found": check the instrument_id matches exactly what is in
  the INSTRUMENTS sheet (case-sensitive, e.g. "ES" not "es").
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")


def _footprint_proxy_row(
    instrument_id: str, source_dataset_id: str
) -> dict[str, object]:
    return {
        "dataset_id":                    f"{instrument_id}_FOOTPRINT_PROXY_1M",
        "dataset_type":                  "derived_trade_metrics_proxy",
        "source_type":                   "canonical",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "incremental",
        "date_col":                      "bar_time",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 "1m",
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "footprint_proxy_1m",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes": (
            f"instrument_id: {instrument_id}\n"
            "metric_type: footprint\n"
            "method: BVC (Bulk Volume Classification) from 1s OHLCV\n"
            "buy_frac=(close-low)/(high-low), fallback 0.5 for doji bars"
        ),
    }


def _cvd_proxy_row(
    instrument_id: str, source_dataset_id: str
) -> dict[str, object]:
    return {
        "dataset_id":                    f"{instrument_id}_CVD_PROXY_1M",
        "dataset_type":                  "derived_trade_metrics_proxy",
        "source_type":                   "canonical",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "incremental",
        "date_col":                      "bar_time",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 "1m",
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "cvd_proxy_1m",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes": (
            f"instrument_id: {instrument_id}\n"
            "metric_type: cvd\n"
            "method: BVC (Bulk Volume Classification) from 1s OHLCV\n"
            "buy_frac=(close-low)/(high-low), fallback 0.5 for doji bars"
        ),
    }


def _add_row_if_missing(wb, row_data: dict[str, object]) -> bool:
    """
    Append row_data to the DATASETS sheet if dataset_id is not already present.
    Returns True if added, False if already present.
    """
    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet in workbook.")

    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]

    if "dataset_id" not in headers:
        raise ValueError("DATASETS sheet has no 'dataset_id' header row.")

    id_col = headers.index("dataset_id") + 1
    existing = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 2)}

    dataset_id = row_data["dataset_id"]
    if dataset_id in existing:
        return False

    ws.append([row_data.get(h) for h in headers])
    return True


def _update_instruments_row(wb, instrument_id: str, fp_proxy_id: str, cvd_proxy_id: str) -> bool:
    """
    Update the INSTRUMENTS row for instrument_id with proxy dataset references.
    Only sets columns that are currently blank.
    Returns True if the row was found, False otherwise.
    """
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet in workbook.")

    ws = wb["INSTRUMENTS"]
    headers = [c.value for c in ws[1]]

    if "instrument_id" not in headers:
        raise ValueError("INSTRUMENTS sheet has no 'instrument_id' header.")

    id_col = headers.index("instrument_id") + 1

    def _col(name: str) -> int | None:
        return (headers.index(name) + 1) if name in headers else None

    fpp_col  = _col("footprint_proxy_dataset_id")
    cvdp_col = _col("cvd_proxy_dataset_id")

    for row_idx in range(2, ws.max_row + 2):
        if ws.cell(row=row_idx, column=id_col).value == instrument_id:
            def _blank(col_idx: int | None) -> bool:
                if col_idx is None:
                    return False
                v = ws.cell(row=row_idx, column=col_idx).value
                return v is None or str(v).strip() == ""

            if fpp_col and _blank(fpp_col):
                ws.cell(row=row_idx, column=fpp_col, value=fp_proxy_id)
            if cvdp_col and _blank(cvdp_col):
                ws.cell(row=row_idx, column=cvdp_col, value=cvd_proxy_id)
            return True

    return False


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Add footprint_proxy_1m and cvd_proxy_1m derived DATASETS rows "
            "to run_config.xlsx for any instrument, and update the INSTRUMENTS row."
        )
    )
    parser.add_argument(
        "--instrument-id",
        required=True,
        metavar="ID",
        help="Instrument ID matching the DATASETS sheet (e.g. ES, NQ).",
    )
    parser.add_argument(
        "--source-dataset-id",
        default=None,
        metavar="SRC",
        help=(
            "Source 1s OHLCV dataset_id (default: DB_{instrument_id}_OHLCV_1S)."
        ),
    )
    args = parser.parse_args()

    instrument_id = args.instrument_id.strip().upper()
    source_dataset_id = (
        args.source_dataset_id.strip()
        if args.source_dataset_id
        else f"DB_{instrument_id}_OHLCV_1S"
    )

    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    fp_proxy_id  = f"{instrument_id}_FOOTPRINT_PROXY_1M"
    cvd_proxy_id = f"{instrument_id}_CVD_PROXY_1M"

    for row_data in [
        _footprint_proxy_row(instrument_id, source_dataset_id),
        _cvd_proxy_row(instrument_id, source_dataset_id),
    ]:
        did = row_data["dataset_id"]
        added = _add_row_if_missing(wb, row_data)
        print(
            f"Added {did} to DATASETS (source: {source_dataset_id})."
            if added
            else f"{did} already in DATASETS, skipping."
        )

    found = _update_instruments_row(wb, instrument_id, fp_proxy_id, cvd_proxy_id)
    print(
        f"Updated INSTRUMENTS {instrument_id} row: "
        f"footprint_proxy_dataset_id={fp_proxy_id}, cvd_proxy_dataset_id={cvd_proxy_id}."
        if found
        else f"WARNING: {instrument_id} row not found in INSTRUMENTS - skipping instrument update."
    )

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
