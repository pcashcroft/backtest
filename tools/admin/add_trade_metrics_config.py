"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds two derived DATASETS rows for any instrument (e.g. ES, NQ) to run_config.xlsx:
    {ID}_FOOTPRINT_1M  (dataset_type=derived_trade_metrics, metric_type=footprint)
    {ID}_CVD_1M        (dataset_type=derived_trade_metrics, metric_type=cvd)
- Re-exports the config snapshot so tools/build/build_derived_trade_metrics.py picks up the change.
- Does NOT modify the INSTRUMENTS sheet (trade-metric datasets are not price sources).

Instrument-agnostic: works for ES, NQ, or any future instrument.

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  # Add ES (uses default source DB_ES_TRADES):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_config.py --instrument-id ES

  # Add NQ (uses default source DB_NQ_TRADES):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_config.py --instrument-id NQ

  # Add with explicit source dataset id:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_trade_metrics_config.py --instrument-id ES --source-dataset-id DB_ES_TRADES

What success looks like:
- Prints "Added {ID}_FOOTPRINT_1M to DATASETS" (or "already exists, skipping").
- Prints "Added {ID}_CVD_1M to DATASETS" (or "already exists, skipping").
- Prints "Config snapshot re-exported".

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")


def _footprint_row(instrument_id: str, source_dataset_id: str) -> dict[str, object]:
    return {
        "dataset_id":                    f"{instrument_id}_FOOTPRINT_1M",
        "dataset_type":                  "derived_trade_metrics",
        "source_type":                   "derived",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "incremental",
        "date_col":                      "bar_time",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 "1m",
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "footprint_base_1m",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes":                         f"instrument_id: {instrument_id}\nmetric_type: footprint",
    }


def _cvd_row(instrument_id: str, source_dataset_id: str) -> dict[str, object]:
    return {
        "dataset_id":                    f"{instrument_id}_CVD_1M",
        "dataset_type":                  "derived_trade_metrics",
        "source_type":                   "derived",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "incremental",
        "date_col":                      "bar_time",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 "1m",
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "cvd_1m",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes":                         f"instrument_id: {instrument_id}\nmetric_type: cvd",
    }


def _add_row_if_missing(wb, row_data: dict[str, object]) -> bool:
    """
    Append row_data to the DATASETS sheet if dataset_id is not already present.
    Returns True if added, False if already present.
    """
    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet in workbook.")

    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]

    if "dataset_id" not in headers:
        raise ValueError("DATASETS sheet has no 'dataset_id' header row.")

    id_col = headers.index("dataset_id") + 1
    existing = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 2)}

    dataset_id = row_data["dataset_id"]
    if dataset_id in existing:
        return False

    ws.append([row_data.get(h) for h in headers])
    return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Add footprint_base_1m and cvd_1m derived DATASETS rows "
            "to run_config.xlsx for any instrument."
        )
    )
    parser.add_argument(
        "--instrument-id",
        required=True,
        metavar="ID",
        help="Instrument ID matching the DATASETS sheet (e.g. ES, NQ).",
    )
    parser.add_argument(
        "--source-dataset-id",
        default=None,
        metavar="SRC",
        help=(
            "Source trades dataset_id (default: DB_{instrument_id}_TRADES)."
        ),
    )
    args = parser.parse_args()

    instrument_id = args.instrument_id.strip().upper()
    source_dataset_id = (
        args.source_dataset_id.strip()
        if args.source_dataset_id
        else f"DB_{instrument_id}_TRADES"
    )

    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    for row_data in [
        _footprint_row(instrument_id, source_dataset_id),
        _cvd_row(instrument_id, source_dataset_id),
    ]:
        did = row_data["dataset_id"]
        added = _add_row_if_missing(wb, row_data)
        print(
            f"Added {did} to DATASETS (source: {source_dataset_id})."
            if added
            else f"{did} already in DATASETS, skipping."
        )

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
