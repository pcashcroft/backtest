"""
INSTRUCTION HEADER
Purpose: Add volume_col and units columns to the INSTRUMENTS sheet if missing.
Inputs: Reads `config/run_config.xlsx`.
Outputs: Updates `config/run_config.xlsx` in place.
How to run: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\migrate_run_config_add_instruments_cols.py`
Success looks like: prints before/after headers and a success message.
Common failures and fixes:
- Permission error: close Excel if the workbook is open and retry.
- Missing workbook: run `tools/admin/make_run_config_xlsx.py`.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    xlsx_path = Path("config/run_config.xlsx")
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = load_workbook(xlsx_path)
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet.")

    ws = wb["INSTRUMENTS"]
    header = [c.value for c in ws[1]]
    before = [h for h in header if h is not None and str(h).strip() != ""]

    additions = []
    for col_name in ["volume_col", "units"]:
        if col_name not in before:
            additions.append(col_name)

    for col_name in additions:
        ws.cell(row=1, column=ws.max_column + 1, value=col_name)

    header_after = [c.value for c in ws[1]]
    after = [h for h in header_after if h is not None and str(h).strip() != ""]

    wb.save(xlsx_path)
    wb.close()

    print("INSTRUMENTS headers (before):", before)
    print("INSTRUMENTS headers (after) :", after)
    print("Migration complete. Added columns:" if additions else "Migration complete. No columns added.", additions)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
