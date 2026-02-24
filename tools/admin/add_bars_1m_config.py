"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds a derived-bars DATASETS row for any instrument (e.g. ES, NQ) to run_config.xlsx.
- Updates the matching INSTRUMENTS row to reference the new dataset as its OHLCV price source.
- Re-exports the config snapshot so tools/build/build_derived_bars_1m.py picks up the change.

Instrument-agnostic: works for ES, NQ, or any future instrument.

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  # Add ES (uses default source DB_ES_OHLCV_1S):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_bars_1m_config.py --instrument-id ES

  # Add NQ (uses default source DB_NQ_OHLCV_1S):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_bars_1m_config.py --instrument-id NQ

  # Add with explicit source dataset id:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_bars_1m_config.py --instrument-id NQ --source-dataset-id DB_NQ_OHLCV_1S

What success looks like:
- Prints "Added {ID}_BARS_1M to DATASETS" (or "already exists, skipping").
- Prints "Updated INSTRUMENTS {ID} row".
- Prints "Config snapshot re-exported".

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
- "INSTRUMENTS row not found": check the instrument_id matches exactly what is in the
  INSTRUMENTS sheet (case-sensitive, e.g. "ES" not "es").
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")


def _build_datasets_row(instrument_id: str, source_dataset_id: str) -> dict[str, object]:
    """Build a DATASETS row dict for the given instrument."""
    return {
        "dataset_id":                     f"{instrument_id}_BARS_1M",
        "dataset_type":                   "derived_bars",
        "source_type":                    "derived",
        "source_path_or_id":              source_dataset_id,
        "update_mode":                    "incremental",
        "date_col":                       "bar_time",
        "timestamp_tz":                   "UTC",
        "bar_frequency":                  "1m",
        "known_time_rule":                "event_time",
        "default_availability_lag_days":  0,
        "canonical_table_name":           "bars_1m",
        "canonical_partition_keys":       "instrument_id,session,date",
        "notes":                          f"instrument_id: {instrument_id}",
    }


def _build_instrument_updates(instrument_id: str) -> dict[str, object]:
    """Build the INSTRUMENTS column updates for the given instrument."""
    return {
        "prices_dataset_id": f"{instrument_id}_BARS_1M",
        "open_col":          "open",
        "high_col":          "high",
        "low_col":           "low",
        "close_col":         "close",
        "volume_col":        "volume",
    }


def _add_datasets_row(wb, instrument_id: str, source_dataset_id: str) -> bool:
    """
    Add a BARS_1M row to the DATASETS sheet for the given instrument.
    Returns True if added, False if the row already exists (skipped).
    """
    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet in workbook.")

    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]

    if "dataset_id" not in headers:
        raise ValueError("DATASETS sheet has no 'dataset_id' header row.")

    dataset_id = f"{instrument_id}_BARS_1M"
    id_col = headers.index("dataset_id") + 1
    existing = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 2)}

    if dataset_id in existing:
        return False

    row_data = _build_datasets_row(instrument_id, source_dataset_id)
    ws.append([row_data.get(h) for h in headers])
    return True


def _update_instruments_row(wb, instrument_id: str) -> bool:
    """
    Update the INSTRUMENTS row matching instrument_id with OHLCV column references.
    Returns True if updated, False if the row was not found.
    """
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet in workbook.")

    ws = wb["INSTRUMENTS"]
    headers = [c.value for c in ws[1]]

    if "instrument_id" not in headers:
        raise ValueError("INSTRUMENTS sheet has no 'instrument_id' header.")

    id_col = headers.index("instrument_id") + 1
    updates = _build_instrument_updates(instrument_id)

    for row_idx in range(2, ws.max_row + 2):
        if ws.cell(row=row_idx, column=id_col).value == instrument_id:
            for field, value in updates.items():
                if field in headers:
                    ws.cell(row=row_idx, column=headers.index(field) + 1, value=value)
            return True

    return False


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Add a derived bars_1m config row to run_config.xlsx for any instrument."
    )
    parser.add_argument(
        "--instrument-id",
        required=True,
        metavar="ID",
        help="Instrument ID as it appears in the INSTRUMENTS sheet (e.g. ES, NQ).",
    )
    parser.add_argument(
        "--source-dataset-id",
        default=None,
        metavar="SRC",
        help=(
            "Source dataset_id for the 1s canonical data "
            "(default: DB_{instrument_id}_OHLCV_1S)."
        ),
    )
    args = parser.parse_args()

    instrument_id = args.instrument_id.strip().upper()
    source_dataset_id = (
        args.source_dataset_id.strip()
        if args.source_dataset_id
        else f"DB_{instrument_id}_OHLCV_1S"
    )

    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    dataset_id = f"{instrument_id}_BARS_1M"
    added = _add_datasets_row(wb, instrument_id, source_dataset_id)
    print(
        f"Added {dataset_id} to DATASETS (source: {source_dataset_id})."
        if added
        else f"{dataset_id} already in DATASETS, skipping."
    )

    updated = _update_instruments_row(wb, instrument_id)
    print(
        f"Updated INSTRUMENTS {instrument_id} row."
        if updated
        else f"WARNING: {instrument_id} row not found in INSTRUMENTS — skipping instrument update."
    )

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
