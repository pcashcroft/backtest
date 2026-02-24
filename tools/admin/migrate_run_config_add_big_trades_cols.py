"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds three big-trade columns to the INSTRUMENTS sheet if they are missing:
    big_trades_dataset_id, big_trades_proxy_dataset_id, big_trades_source_mode
- Sets default values for any instrument row whose big-trade datasets already
  exist in the DATASETS sheet (e.g. ES_BIG_TRADES, ES_BIG_TRADES_PROXY for ES).
- Re-exports the config snapshot.

This is a one-time migration. Running it again is safe (idempotent - only adds
missing columns; only sets values that are currently blank).

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\migrate_run_config_add_big_trades_cols.py

What success looks like:
- Prints the columns added (or "already present").
- Prints each instrument row updated with its default values.
- Prints "Config snapshot re-exported."

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")

NEW_COLS = [
    "big_trades_dataset_id",
    "big_trades_proxy_dataset_id",
    "big_trades_source_mode",
]

DEFAULT_BIG_TRADES_SOURCE_MODE = "real_then_proxy"


def _add_missing_columns(ws) -> list[str]:
    """Append any missing NEW_COLS to the header row. Returns list of added cols."""
    header = [c.value for c in ws[1]]
    added = []
    for col_name in NEW_COLS:
        if col_name not in header:
            # Insert before the existing 'notes' column if present, else append
            if "notes" in header:
                notes_idx = header.index("notes") + 1  # 1-based
                ws.insert_cols(notes_idx)
                ws.cell(row=1, column=notes_idx, value=col_name)
                # refresh header list for subsequent insertions
                header = [c.value for c in ws[1]]
            else:
                ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            added.append(col_name)
    return added


def _get_dataset_ids(wb) -> set[str]:
    """Return the set of dataset_ids from the DATASETS sheet."""
    if "DATASETS" not in wb.sheetnames:
        return set()
    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]
    if "dataset_id" not in headers:
        return set()
    id_col = headers.index("dataset_id") + 1
    return {
        ws.cell(row=r, column=id_col).value
        for r in range(2, ws.max_row + 2)
        if ws.cell(row=r, column=id_col).value
    }


def _set_instrument_defaults(ws, dataset_ids: set[str]) -> list[str]:
    """
    For each instrument row, set big-trade defaults if the columns are blank.
    Only sets values if the referenced datasets actually exist in DATASETS.
    Returns a list of summary strings describing what was set.
    """
    header = [c.value for c in ws[1]]

    def col(name: str) -> int | None:
        return (header.index(name) + 1) if name in header else None

    id_col = col("instrument_id")
    if id_col is None:
        return []

    bt_col   = col("big_trades_dataset_id")
    btp_col  = col("big_trades_proxy_dataset_id")
    mode_col = col("big_trades_source_mode")

    updated = []
    for row_idx in range(2, ws.max_row + 2):
        iid = ws.cell(row=row_idx, column=id_col).value
        if not iid:
            continue

        changes = {}

        bt_id  = f"{iid}_BIG_TRADES"
        btp_id = f"{iid}_BIG_TRADES_PROXY"

        def _blank(col_idx: int | None) -> bool:
            if col_idx is None:
                return False
            v = ws.cell(row=row_idx, column=col_idx).value
            return v is None or str(v).strip() == ""

        def _set(col_idx: int | None, value: str) -> None:
            if col_idx is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

        if bt_id in dataset_ids and _blank(bt_col):
            _set(bt_col, bt_id)
            changes["big_trades_dataset_id"] = bt_id

        if btp_id in dataset_ids and _blank(btp_col):
            _set(btp_col, btp_id)
            changes["big_trades_proxy_dataset_id"] = btp_id

        has_any = bt_id in dataset_ids or btp_id in dataset_ids
        if has_any and _blank(mode_col):
            _set(mode_col, DEFAULT_BIG_TRADES_SOURCE_MODE)
            changes["big_trades_source_mode"] = DEFAULT_BIG_TRADES_SOURCE_MODE

        if changes:
            updated.append(f"  {iid}: {changes}")

    return updated


def main() -> int:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet.")

    ws = wb["INSTRUMENTS"]

    # Step 1: add missing columns
    added = _add_missing_columns(ws)
    if added:
        print(f"Added columns to INSTRUMENTS: {added}")
    else:
        print(f"All big-trade columns already present: {NEW_COLS}")

    # Step 2: collect known dataset IDs
    dataset_ids = _get_dataset_ids(wb)

    # Step 3: set defaults for instrument rows
    updated = _set_instrument_defaults(ws, dataset_ids)
    if updated:
        print("Set big-trade defaults:")
        for line in updated:
            print(line)
    else:
        print("No instrument rows updated (all already set or no matching datasets found).")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
