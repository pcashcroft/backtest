"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds five threshold columns to the DATASETS sheet if they are missing:
    threshold_method    — fixed_count / rolling_pct / z_score
    threshold_min_size  — minimum contract count (for fixed_count)
    threshold_pct       — percentile 0-100 (for rolling_pct, e.g. 99.0)
    threshold_z         — z-score multiplier (for z_score, e.g. 2.5)
    threshold_window_days — lookback window in calendar days (for rolling_pct + z_score)
- Sets defaults for ES_BIG_TRADES and ES_BIG_TRADES_PROXY if those rows exist:
    ES_BIG_TRADES:       threshold_method=fixed_count, threshold_min_size=50
    ES_BIG_TRADES_PROXY: threshold_method=fixed_count, threshold_min_size=100
- Cleans the stale threshold_method / min_size values out of the notes field for
  big_trades rows (leaves instrument_id: ES intact).
- Re-exports the config snapshot.

This is a one-time migration. Running it again is safe (idempotent — only adds
missing columns; only sets values that are currently blank).

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\migrate_run_config_add_threshold_cols.py

What success looks like:
- Prints the columns added (or "already present").
- Prints each DATASETS row updated with default values.
- Prints "Config snapshot re-exported."

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
"""

from __future__ import annotations

import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")

NEW_COLS = [
    "threshold_method",
    "threshold_min_size",
    "threshold_pct",
    "threshold_z",
    "threshold_window_days",
]

# Defaults keyed by dataset_id suffix pattern
BIG_TRADES_DEFAULTS = {
    "threshold_method":        "fixed_count",
    "threshold_min_size":      50,
    "threshold_pct":           None,
    "threshold_z":             None,
    "threshold_window_days":   None,
}
BIG_TRADES_PROXY_DEFAULTS = {
    "threshold_method":        "fixed_count",
    "threshold_min_size":      100,
    "threshold_pct":           None,
    "threshold_z":             None,
    "threshold_window_days":   None,
}

# Notes lines to strip out of big_trades rows (now promoted to columns)
_NOTES_STRIP = re.compile(
    r"^\s*(threshold_method|min_size)\s*:.*$", re.MULTILINE
)


def _add_missing_columns(ws) -> list[str]:
    """Append any missing NEW_COLS to the DATASETS header row before 'notes'.
    Returns list of added cols."""
    header = [c.value for c in ws[1]]
    added = []
    for col_name in NEW_COLS:
        if col_name not in header:
            if "notes" in header:
                notes_idx = header.index("notes") + 1  # 1-based
                ws.insert_cols(notes_idx)
                ws.cell(row=1, column=notes_idx, value=col_name)
                header = [c.value for c in ws[1]]
            else:
                ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            added.append(col_name)
    return added


def _clean_notes(notes_val: str | None) -> str:
    """Strip threshold_method / min_size lines from notes (now in columns)."""
    if not notes_val:
        return notes_val or ""
    cleaned = _NOTES_STRIP.sub("", notes_val or "")
    # Collapse multiple blank lines
    cleaned = re.sub(r"\n{2,}", "\n", cleaned).strip()
    return cleaned


def _set_dataset_defaults(ws) -> list[str]:
    """
    For each DATASETS row of type big_trades / big_trades_proxy, set threshold
    column defaults if currently blank and clean the notes field.
    Returns summary lines.
    """
    header = [c.value for c in ws[1]]

    def col(name: str) -> int | None:
        return (header.index(name) + 1) if name in header else None

    id_col   = col("dataset_id")
    type_col = col("dataset_type")
    notes_col = col("notes")
    if id_col is None:
        return []

    def _blank(col_idx: int | None, row: int) -> bool:
        if col_idx is None:
            return False
        v = ws.cell(row=row, column=col_idx).value
        return v is None or str(v).strip() == ""

    def _set(col_idx: int | None, row: int, value) -> None:
        if col_idx is not None and value is not None:
            ws.cell(row=row, column=col_idx, value=value)

    updated = []
    for row_idx in range(2, ws.max_row + 2):
        did = ws.cell(row=row_idx, column=id_col).value
        if not did:
            continue
        dtype = ws.cell(row=row_idx, column=type_col).value if type_col else None
        if dtype not in ("big_trades", "big_trades_proxy"):
            continue

        defaults = (
            BIG_TRADES_DEFAULTS if dtype == "big_trades"
            else BIG_TRADES_PROXY_DEFAULTS
        )
        changes = {}
        for field, default_val in defaults.items():
            c = col(field)
            if _blank(c, row_idx):
                _set(c, row_idx, default_val)
                if default_val is not None:
                    changes[field] = default_val

        # Clean stale threshold info out of notes
        if notes_col is not None:
            raw = ws.cell(row=row_idx, column=notes_col).value
            cleaned = _clean_notes(raw)
            if cleaned != (raw or ""):
                ws.cell(row=row_idx, column=notes_col, value=cleaned or None)
                changes["notes"] = "(threshold lines removed)"

        if changes:
            updated.append(f"  {did}: {changes}")

    return updated


def main() -> int:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet.")

    ws = wb["DATASETS"]

    # Step 1: add missing columns
    added = _add_missing_columns(ws)
    if added:
        print(f"Added columns to DATASETS: {added}")
    else:
        print(f"All threshold columns already present: {NEW_COLS}")

    # Step 2: set defaults for big_trades rows
    updated = _set_dataset_defaults(ws)
    if updated:
        print("Set threshold defaults:")
        for line in updated:
            print(line)
    else:
        print("No DATASETS rows updated (all already set or no big_trades rows found).")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
