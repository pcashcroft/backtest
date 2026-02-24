"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds two big-trade DATASETS rows for any instrument (e.g. ES, NQ) to run_config.xlsx:
    {ID}_BIG_TRADES        (dataset_type=big_trades,       source=real trades)
    {ID}_BIG_TRADES_PROXY  (dataset_type=big_trades_proxy, source=1s OHLCV)
- Updates the matching INSTRUMENTS row to reference the new datasets and set
  big_trades_source_mode = real_then_proxy.
- Re-exports the config snapshot so the big_trades compute function picks up the change.

Threshold config is stored in DATASETS.notes as KEY: VALUE pairs, defaulting to
fixed_count with instrument-specific min_size values. To switch to rolling_pct or
z_score, edit the notes field in the DATASETS sheet and re-export the snapshot.

Supported threshold_method values and their required notes keys:
  fixed_count  ->  min_size: <int>
  rolling_pct  ->  pct: <float 0-100>,  window_days: <int>
  z_score      ->  z_threshold: <float>, window_days: <int>

Instrument-agnostic: works for ES, NQ, or any future instrument.

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  # Add ES (default min_size=50 real, 100 proxy):
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_big_trades_config.py --instrument-id ES

  # Add NQ:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_big_trades_config.py --instrument-id NQ

  # Add ES with custom thresholds:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\admin\\add_big_trades_config.py --instrument-id ES --min-size 100 --proxy-min-size 200

What success looks like:
- Prints "Added {ID}_BIG_TRADES to DATASETS" (or "already exists, skipping").
- Prints "Added {ID}_BIG_TRADES_PROXY to DATASETS" (or "already exists, skipping").
- Prints "Updated INSTRUMENTS {ID} row."
- Prints "Config snapshot re-exported."

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/admin/make_run_config_xlsx.py first.
- "INSTRUMENTS row not found": check the instrument_id matches exactly what is in the
  INSTRUMENTS sheet (case-sensitive, e.g. "ES" not "es").
- "big_trades_dataset_id column not found": run
  tools/admin/migrate_run_config_add_big_trades_cols.py first.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")


def _big_trades_row(
    instrument_id: str,
    source_dataset_id: str,
    min_size: int,
) -> dict[str, object]:
    notes = (
        f"instrument_id: {instrument_id}\n"
        f"threshold_method: fixed_count\n"
        f"min_size: {min_size}"
    )
    return {
        "dataset_id":                    f"{instrument_id}_BIG_TRADES",
        "dataset_type":                  "big_trades",
        "source_type":                   "canonical",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "on_the_fly",
        "date_col":                      "ts_event",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 None,
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "big_trade_events",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes":                         notes,
    }


def _big_trades_proxy_row(
    instrument_id: str,
    source_dataset_id: str,
    proxy_min_size: int,
) -> dict[str, object]:
    notes = (
        f"instrument_id: {instrument_id}\n"
        f"threshold_method: fixed_count\n"
        f"min_size: {proxy_min_size}"
    )
    return {
        "dataset_id":                    f"{instrument_id}_BIG_TRADES_PROXY",
        "dataset_type":                  "big_trades_proxy",
        "source_type":                   "canonical",
        "source_path_or_id":             source_dataset_id,
        "update_mode":                   "on_the_fly",
        "date_col":                      "ts_event",
        "timestamp_tz":                  "UTC",
        "bar_frequency":                 "1s",
        "known_time_rule":               "event_time",
        "default_availability_lag_days": 0,
        "canonical_table_name":          "big_trade_events_proxy",
        "canonical_partition_keys":      "instrument_id,session,date",
        "notes":                         notes,
    }


def _add_row_if_missing(wb, row_data: dict[str, object]) -> bool:
    """
    Append row_data to the DATASETS sheet if dataset_id is not already present.
    Returns True if added, False if already present.
    """
    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet in workbook.")

    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]

    if "dataset_id" not in headers:
        raise ValueError("DATASETS sheet has no 'dataset_id' header row.")

    id_col = headers.index("dataset_id") + 1
    existing = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 2)}

    dataset_id = row_data["dataset_id"]
    if dataset_id in existing:
        return False

    ws.append([row_data.get(h) for h in headers])
    return True


def _update_instruments_row(wb, instrument_id: str) -> bool:
    """
    Update the INSTRUMENTS row matching instrument_id with big-trade dataset references.
    Returns True if updated, False if row not found.
    """
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet in workbook.")

    ws = wb["INSTRUMENTS"]
    headers = [c.value for c in ws[1]]

    if "instrument_id" not in headers:
        raise ValueError("INSTRUMENTS sheet has no 'instrument_id' header.")

    updates = {
        "big_trades_dataset_id":       f"{instrument_id}_BIG_TRADES",
        "big_trades_proxy_dataset_id": f"{instrument_id}_BIG_TRADES_PROXY",
        "big_trades_source_mode":      "real_then_proxy",
    }

    id_col = headers.index("instrument_id") + 1
    for row_idx in range(2, ws.max_row + 2):
        if ws.cell(row=row_idx, column=id_col).value == instrument_id:
            for field, value in updates.items():
                if field in headers:
                    ws.cell(row=row_idx, column=headers.index(field) + 1, value=value)
                else:
                    print(
                        f"  WARNING: column '{field}' not found in INSTRUMENTS — "
                        "run migrate_run_config_add_big_trades_cols.py first."
                    )
            return True

    return False


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Add big_trade_events and big_trade_events_proxy DATASETS rows "
            "to run_config.xlsx for any instrument."
        )
    )
    parser.add_argument(
        "--instrument-id",
        required=True,
        metavar="ID",
        help="Instrument ID matching the INSTRUMENTS sheet (e.g. ES, NQ).",
    )
    parser.add_argument(
        "--source-dataset-id",
        default=None,
        metavar="SRC",
        help="Source trades dataset_id for real data (default: DB_{instrument_id}_TRADES).",
    )
    parser.add_argument(
        "--proxy-source-dataset-id",
        default=None,
        metavar="PSRC",
        help="Source dataset_id for proxy (default: DB_{instrument_id}_OHLCV_1S).",
    )
    parser.add_argument(
        "--min-size",
        type=int,
        default=50,
        metavar="N",
        help="Default fixed_count min_size for real trades (default: 50).",
    )
    parser.add_argument(
        "--proxy-min-size",
        type=int,
        default=100,
        metavar="N",
        help="Default fixed_count min_size for proxy (default: 100).",
    )
    args = parser.parse_args()

    instrument_id = args.instrument_id.strip().upper()
    source_dataset_id = (
        args.source_dataset_id.strip()
        if args.source_dataset_id
        else f"DB_{instrument_id}_TRADES"
    )
    proxy_source_dataset_id = (
        args.proxy_source_dataset_id.strip()
        if args.proxy_source_dataset_id
        else f"DB_{instrument_id}_OHLCV_1S"
    )

    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/admin/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    # Add DATASETS rows
    for row_data in [
        _big_trades_row(instrument_id, source_dataset_id, args.min_size),
        _big_trades_proxy_row(instrument_id, proxy_source_dataset_id, args.proxy_min_size),
    ]:
        did = row_data["dataset_id"]
        added = _add_row_if_missing(wb, row_data)
        print(
            f"Added {did} to DATASETS."
            if added
            else f"{did} already in DATASETS, skipping."
        )

    # Update INSTRUMENTS row
    updated = _update_instruments_row(wb, instrument_id)
    print(
        f"Updated INSTRUMENTS {instrument_id} row."
        if updated
        else f"WARNING: {instrument_id} row not found in INSTRUMENTS — skipping instrument update."
    )

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/admin/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
