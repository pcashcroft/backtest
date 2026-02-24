"""
INSTRUCTION HEADER

What this script does (plain English):
- Adds five metric-source columns to the INSTRUMENTS sheet if they are missing:
    footprint_dataset_id, footprint_proxy_dataset_id,
    cvd_dataset_id, cvd_proxy_dataset_id, metric_source_mode
- Sets default values for the ES row (and any other instrument rows that have
  a matching FOOTPRINT_1M / CVD_1M dataset in the DATASETS sheet).
- Re-exports the config snapshot.

This is a one-time migration. Running it again is safe (idempotent - only adds
missing columns; only sets values that are currently blank).

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
  C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\migrate_run_config_add_metric_source_cols.py

What success looks like:
- Prints the columns added (or "already present").
- Prints each instrument row updated with its default values.
- Prints "Config snapshot re-exported."

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/make_run_config_xlsx.py first.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")

NEW_COLS = [
    "footprint_dataset_id",
    "footprint_proxy_dataset_id",
    "cvd_dataset_id",
    "cvd_proxy_dataset_id",
    "metric_source_mode",
]

DEFAULT_METRIC_SOURCE_MODE = "real_then_proxy"


def _add_missing_columns(ws) -> list[str]:
    """Append any missing NEW_COLS to the header row. Returns list of added cols."""
    header = [c.value for c in ws[1]]
    added = []
    for col_name in NEW_COLS:
        if col_name not in header:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            added.append(col_name)
    return added


def _get_dataset_ids(wb) -> set[str]:
    """Return the set of dataset_ids from the DATASETS sheet."""
    if "DATASETS" not in wb.sheetnames:
        return set()
    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]
    if "dataset_id" not in headers:
        return set()
    id_col = headers.index("dataset_id") + 1
    return {
        ws.cell(row=r, column=id_col).value
        for r in range(2, ws.max_row + 2)
        if ws.cell(row=r, column=id_col).value
    }


def _set_instrument_defaults(ws, dataset_ids: set[str]) -> list[str]:
    """
    For each instrument row, set metric-source defaults if the columns are blank.
    Only sets values if the referenced datasets actually exist in DATASETS.
    Returns a list of summary strings describing what was set.
    """
    header = [c.value for c in ws[1]]

    def col(name: str) -> int | None:
        return (header.index(name) + 1) if name in header else None

    id_col = col("instrument_id")
    if id_col is None:
        return []

    fp_col   = col("footprint_dataset_id")
    fpp_col  = col("footprint_proxy_dataset_id")
    cvd_col  = col("cvd_dataset_id")
    cvdp_col = col("cvd_proxy_dataset_id")
    mode_col = col("metric_source_mode")

    updated = []
    for row_idx in range(2, ws.max_row + 2):
        iid = ws.cell(row=row_idx, column=id_col).value
        if not iid:
            continue

        changes = {}

        fp_id   = f"{iid}_FOOTPRINT_1M"
        fpp_id  = f"{iid}_FOOTPRINT_PROXY_1M"
        cvd_id  = f"{iid}_CVD_1M"
        cvdp_id = f"{iid}_CVD_PROXY_1M"

        def _blank(col_idx: int | None) -> bool:
            if col_idx is None:
                return False
            v = ws.cell(row=row_idx, column=col_idx).value
            return v is None or str(v).strip() == ""

        def _set(col_idx: int | None, value: str) -> None:
            if col_idx is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

        if fp_id in dataset_ids and _blank(fp_col):
            _set(fp_col, fp_id)
            changes["footprint_dataset_id"] = fp_id

        if fpp_id in dataset_ids and _blank(fpp_col):
            _set(fpp_col, fpp_id)
            changes["footprint_proxy_dataset_id"] = fpp_id

        if cvd_id in dataset_ids and _blank(cvd_col):
            _set(cvd_col, cvd_id)
            changes["cvd_dataset_id"] = cvd_id

        if cvdp_id in dataset_ids and _blank(cvdp_col):
            _set(cvdp_col, cvdp_id)
            changes["cvd_proxy_dataset_id"] = cvdp_id

        # Set mode if any metric dataset exists for this instrument and mode is blank
        has_any = any(
            d in dataset_ids for d in [fp_id, fpp_id, cvd_id, cvdp_id]
        )
        if has_any and _blank(mode_col):
            _set(mode_col, DEFAULT_METRIC_SOURCE_MODE)
            changes["metric_source_mode"] = DEFAULT_METRIC_SOURCE_MODE

        if changes:
            updated.append(f"  {iid}: {changes}")

    return updated


def main() -> int:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet.")

    ws = wb["INSTRUMENTS"]

    # Step 1: add missing columns
    added = _add_missing_columns(ws)
    if added:
        print(f"Added columns to INSTRUMENTS: {added}")
    else:
        print(f"All metric-source columns already present: {NEW_COLS}")

    # Step 2: collect known dataset IDs
    dataset_ids = _get_dataset_ids(wb)

    # Step 3: set defaults for instrument rows
    updated = _set_instrument_defaults(ws, dataset_ids)
    if updated:
        print("Set metric-source defaults:")
        for line in updated:
            print(line)
    else:
        print("No instrument rows updated (all already set or no matching datasets found).")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
