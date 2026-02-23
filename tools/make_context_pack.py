"""
INSTRUCTION HEADER
Purpose: Generate a full project context pack for ChatGPT/Codex.
Inputs: Reads repo files, key docs, and config snapshots if present.
Outputs: Writes `context_pack.md` and prints it to the terminal.
How to run: `pybt tools/make_context_pack.py`
Also: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\make_context_pack.py`
Success looks like: the full context prints to the terminal and `context_pack.md` exists.
Common failures and fixes:
- Git not installed: script still runs, but git metadata may show as `(no git)`.
- Unicode/binary files: remove binary files from repo root and retry.
"""

from __future__ import annotations

from pathlib import Path
import datetime as dt
import subprocess
import sys


def _repo_root() -> Path:
    """Return the repository root folder based on this file location."""
    return Path(__file__).resolve().parents[1]


def _run_git(args: list[str]) -> str:
    """Run a git command and return its output, or empty string on failure."""
    try:
        out = subprocess.check_output(["git", *args], stderr=subprocess.STDOUT)
        return out.decode("utf-8", errors="replace").strip()
    except Exception:
        return ""


def _folder_tree(root: Path, max_depth: int = 3) -> list[str]:
    """Return a folder tree up to the requested depth."""
    excluded = {
        ".git",
        "__pycache__",
        ".venv",
        "venv",
        ".ipynb_checkpoints",
        ".pytest_cache",
        ".mypy_cache",
        ".ruff_cache",
    }
    lines: list[str] = []
    root = root.resolve()

    def walk(path: Path, depth: int) -> None:
        if depth > max_depth:
            return
        entries = sorted(path.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
        for entry in entries:
            if entry.name in excluded:
                continue
            rel = entry.relative_to(root)
            indent = "  " * (depth - 1)
            lines.append(f"{indent}{rel.as_posix()}")
            if entry.is_dir():
                walk(entry, depth + 1)

    walk(root, 1)
    return lines


def _read_text(path: Path) -> str:
    """Read a UTF-8 text file from disk (replace errors)."""
    # Be resilient to non-UTF8 text files; never crash the context pack.
    return path.read_text(encoding="utf-8", errors="replace")


def _git_changed_files_last_commits(n: int = 5) -> list[Path]:
    """Return unique files changed in the last N commits."""
    out = _run_git(["log", f"-n{n}", "--name-only", "--pretty=format:"])
    if not out:
        return []
    files = []
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        files.append(line)
    uniq = []
    seen = set()
    for f in files:
        if f not in seen:
            seen.add(f)
            uniq.append(Path(f))
    return uniq


def _is_binary_extension(path: Path) -> bool:
    """Return True if a file is likely binary and should be skipped."""
    # Keep this conservative: if it's likely binary, don't attempt read_text().
    return path.suffix.lower() in {
        ".xlsx",
        ".xls",
        ".xlsm",
        ".parquet",
        ".duckdb",
        ".db",
        ".sqlite",
        ".png",
        ".jpg",
        ".jpeg",
        ".gif",
        ".webp",
        ".pdf",
        ".zip",
        ".7z",
        ".tar",
        ".gz",
        ".bz2",
        ".xz",
    }


def _summarize_xlsx_headers(path: Path) -> str:
    """Return a summary of sheet headers for an XLSX file."""
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        return f"(cannot summarize xlsx; openpyxl missing) {e!r}"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return f"(cannot open xlsx) {e!r}"

    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Header row is expected in row 1.
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = []
        if row:
            for v in row:
                if v is None:
                    headers.append("")
                else:
                    headers.append(str(v))
        lines.append(f"- {sheet_name}: {', '.join(headers)}")
    return "\n".join(lines)


def _include_file(path: Path) -> str:
    """Return full file text or a trimmed version if it is large."""
    if not path.exists() or not path.is_file():
        return f"(missing) {path.as_posix()}"

    if _is_binary_extension(path):
        if path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}:
            return _summarize_xlsx_headers(path)
        return f"(binary file skipped) {path.as_posix()}"

    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    if len(lines) < 300:
        body = "\n".join(lines)
    else:
        top = "\n".join(lines[:120])
        bottom = "\n".join(lines[-60:])
        body = f"{top}\n...\n{bottom}"
    return body


def main() -> int:
    """Generate `context_pack.md` and print it to the terminal."""
    repo_root = _repo_root()
    ts = dt.datetime.now().isoformat(timespec="seconds")
    branch = _run_git(["rev-parse", "--abbrev-ref", "HEAD"]) or "(no git)"
    head = _run_git(["rev-parse", "HEAD"]) or "(no git)"
    last5 = _run_git(["log", "-n5", "--oneline"]) or "(no git)"

    parts: list[str] = []
    parts.append("# Context Pack")
    parts.append("")
    parts.append("## Repo Metadata")
    parts.append(f"- Timestamp: {ts}")
    parts.append(f"- Branch: {branch}")
    parts.append(f"- HEAD: {head}")
    parts.append("### Last 5 Commits")
    parts.append("```text")
    parts.append(last5)
    parts.append("```")
    parts.append("")

    parts.append("## Folder Tree (2-3 levels)")
    parts.append("```text")
    parts.extend(_folder_tree(repo_root, max_depth=3))
    parts.append("```")
    parts.append("")

    parts.append("## Key Docs")
    key_docs = [
        repo_root / "design" / "SPEC.md",
        repo_root / "design" / "WORKFLOW.md",
        repo_root / "design" / "ROADMAP.md",
        repo_root / "instructions" / "00_setup.md",
        repo_root / "instructions" / "01_daily_workflow.md",
    ]
    for doc in key_docs:
        parts.append(f"### {doc.relative_to(repo_root).as_posix()}")
        parts.append("```text")
        parts.append(_read_text(doc))
        parts.append("```")
        parts.append("")

    parts.append("## Config")
    latest_snapshot = repo_root / "config" / "exports" / "config_snapshot_latest.json"
    if latest_snapshot.exists():
        parts.append("### config/exports/config_snapshot_latest.json")
        parts.append("```text")
        parts.append(_read_text(latest_snapshot))
        parts.append("```")
    else:
        sys.path.insert(0, str(repo_root / "src"))
        from backtest.config.schema import HEADERS

        parts.append("### Workbook Schema")
        for sheet, headers in HEADERS.items():
            parts.append(f"- {sheet}: {', '.join(headers)}")
    parts.append("")

    parts.append("## Code Focus")
    changed = _git_changed_files_last_commits(5)
    if not changed:
        parts.append("(no git history or no changed files)")
    else:
        for path in changed:
            full_path = repo_root / path
            parts.append(f"### {path.as_posix()}")
            parts.append("```text")
            parts.append(_include_file(full_path))
            parts.append("```")
    parts.append("")

    output_path = repo_root / "context_pack.md"
    output_path.write_text("\n".join(parts), encoding="utf-8")
    print(output_path.read_text(encoding="utf-8", errors="replace"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
