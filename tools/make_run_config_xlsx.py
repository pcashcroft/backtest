"""
INSTRUCTION HEADER
Purpose: Create the Excel control-plane workbook with required sheets and headers.
Inputs: Reads schema from `src/backtest/config/schema.py`.
Outputs: Writes `config/run_config.xlsx`.
How to run: `pybt tools/make_run_config_xlsx.py`
Also: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\make_run_config_xlsx.py`
Success looks like: console prints `Created workbook: ...config\\run_config.xlsx`.
Common failures and fixes:
- Module not found (openpyxl): run `pybt -m pip install openpyxl`.
- Wrong working folder: run from the repo root.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _repo_root() -> Path:
    """Return the repository root folder based on this file location."""
    return Path(__file__).resolve().parents[1]


def _load_headers() -> dict[str, list[str]]:
    """Load the Excel header schema from `src/backtest/config/schema.py`."""
    import sys

    repo_root = _repo_root()
    sys.path.insert(0, str(repo_root / "src"))
    from backtest.config.schema import HEADERS

    return HEADERS


def main() -> int:
    """Generate the workbook and write it to `config/run_config.xlsx`."""
    headers = _load_headers()
    repo_root = _repo_root()
    output_path = repo_root / "config" / "run_config.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)

    for sheet_name, cols in headers.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    code_root = str(repo_root)
    export_dir = str(repo_root / "config" / "exports")

    paths_row = [
        "DEFAULT",
        1,
        code_root,
        "E:\\BacktestData",
        "E:\\BacktestData\\duckdb\\research.duckdb",
        "E:\\BacktestData\\raw",
        "E:\\BacktestData\\canonical",
        "E:\\BacktestData\\features_cache",
        "E:\\BacktestData\\runs",
        "E:\\BacktestData\\logs",
        export_dir,
        "Default local profile; edit in Excel only",
    ]
    ws_paths = wb["PATHS"]
    ws_paths.append(paths_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
