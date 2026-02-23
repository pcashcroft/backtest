"""
INSTRUCTION HEADER
Purpose: Export the Excel control plane to JSON snapshots.
Inputs: Reads `config/run_config.xlsx`.
Outputs: Writes `config/exports/config_snapshot_<YYYYMMDD_HHMMSS>.json`
and `config/exports/config_snapshot_latest.json`.
How to run: `pybt tools/export_config_snapshot.py`
Also: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\export_config_snapshot.py`
Success looks like: printed paths for the timestamped snapshot and latest snapshot.
Common failures and fixes:
- Module not found (openpyxl or orjson): run `pybt -m pip install openpyxl orjson`.
- Missing workbook: run `pybt tools/make_run_config_xlsx.py`.
"""

from __future__ import annotations

from pathlib import Path
import datetime as dt
import json

from openpyxl import load_workbook


def _repo_root() -> Path:
    """Return the repository root folder based on this file location."""
    return Path(__file__).resolve().parents[1]


def main() -> int:
    """Export snapshots using the schema from the workbook itself."""
    repo_root = _repo_root()
    xlsx_path = repo_root / "config" / "run_config.xlsx"
    exports_dir = repo_root / "config" / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    latest_path = exports_dir / "config_snapshot_latest.json"
    stamped_path = exports_dir / f"config_snapshot_{ts}.json"

    wb = load_workbook(xlsx_path)

    schema: dict[str, list[str]] = {}
    sheets: dict[str, list[dict[str, object]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [c.value for c in ws[1]]
        header = [h for h in header_row if h is not None and str(h).strip() != ""]
        if not header:
            raise ValueError(f"Missing header row in sheet: {sheet_name}")

        def _is_blank_row(values: list[object]) -> bool:
            for v in values:
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                return False
            return True

        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if _is_blank_row(list(row)):
                continue
            record = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
            rows.append(record)

        schema[sheet_name] = header
        sheets[sheet_name] = rows

    payload = {
        "exported_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": str(xlsx_path),
        "schema": schema,
        "sheets": sheets,
    }

    stamped_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    latest_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"Wrote snapshot: {stamped_path}")
    print(f"Wrote latest : {latest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
