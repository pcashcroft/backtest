"""
INSTRUCTION HEADER
Purpose: Verify the Excel control-plane workbook matches the expected schema.
Inputs: Reads `config/run_config.xlsx` and schema in `src/backtest/config/schema.py`.
Outputs: None (prints results, exits non-zero on failure).
How to run: `pybt tools/verify_run_config_xlsx.py`
Also: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\verify_run_config_xlsx.py`
Success looks like: `Workbook verification passed.`
Common failures and fixes:
- Module not found (openpyxl): run `pybt -m pip install openpyxl`.
- Header mismatch: regenerate the workbook with `pybt tools/make_run_config_xlsx.py`.
"""

from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook


def _repo_root() -> Path:
    """Return the repository root folder based on this file location."""
    return Path(__file__).resolve().parents[1]


def _load_headers() -> dict[str, list[str]]:
    """Load the Excel header schema from `src/backtest/config/schema.py`."""
    repo_root = _repo_root()
    sys.path.insert(0, str(repo_root / "src"))
    from backtest.config.schema import HEADERS

    return HEADERS


def main() -> int:
    """Verify workbook sheets and headers against the schema."""
    headers = _load_headers()
    repo_root = _repo_root()
    xlsx_path = repo_root / "config" / "run_config.xlsx"

    if not xlsx_path.exists():
        print(f"Missing workbook: {xlsx_path}")
        return 1

    wb = load_workbook(xlsx_path)
    expected_sheets = list(headers.keys())
    actual_sheets = wb.sheetnames

    missing = [s for s in expected_sheets if s not in actual_sheets]
    extra = [s for s in actual_sheets if s not in expected_sheets]

    failed = False
    if missing:
        print("Missing sheets:", ", ".join(missing))
        failed = True
    if extra:
        print("Unexpected sheets:", ", ".join(extra))
        failed = True

    for sheet_name, required_cols in headers.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        actual_cols = [c.value for c in ws[1] if c.value is not None and str(c.value).strip() != ""]
        missing_cols = [c for c in required_cols if c not in actual_cols]
        if missing_cols:
            failed = True
            print(f"Missing required columns in {sheet_name}: {missing_cols}")

    if failed:
        return 1

    print("Workbook verification passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
