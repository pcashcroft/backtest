"""
INSTRUCTION HEADER
Purpose: Populate INSTRUMENTS rows for macro instruments and add ES placeholder from Macro_Instruments.xlsx.
Inputs: `config/run_config.xlsx`, `E:/BacktestData/raw/Macro_Instruments.xlsx`.
Outputs: Updates `config/run_config.xlsx` in place.
How to run: `C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\update_instruments_from_macro_workbook.py`
Success looks like: prints counts of updated/appended instrument rows.
Common failures and fixes:
- Permission error: close Excel if the workbook is open and retry.
- Missing workbook: verify paths and SSD mount.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


MACRO_PATH = Path(r"E:\BacktestData\raw\Macro_Instruments.xlsx")
CONFIG_PATH = Path("config/run_config.xlsx")

SHEET_TO_INSTRUMENT = {
    "Xover": "XOVER",
    "CDX_HY": "CDX_HY",
    "VIX": "VIX",
    "V2X": "V2X",
    "SPX": "SPX",
    "SX5E": "SX5E",
    "USGG2YR": "USGG2YR",
    "USGG10YR": "USGG10YR",
    "SPX_AD": "SPX_AD",
    "SPX_PC": "SPX_PC",
}

UNITS = {
    "XOVER": "spread_bp",
    "CDX_HY": "spread_bp",
    "SPX": "index_points",
    "SX5E": "index_points",
    "VIX": "vol_index",
    "V2X": "vol_index",
    "USGG2YR": "yield_pct",
    "USGG10YR": "yield_pct",
    "SPX_AD": "net_advancers",
    "SPX_PC": "ratio",
}


def _load_macro_headers(path: Path) -> dict[str, list[str]]:
    if not path.exists():
        raise FileNotFoundError(path)
    xls = pd.ExcelFile(path, engine="openpyxl")
    headers: dict[str, list[str]] = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", nrows=0)
        headers[sheet] = list(df.columns)
    return headers


def _build_macro_row(
    instrument_id: str,
    prefix: str,
    has_volume: bool,
    last_only: bool,
) -> dict[str, object]:
    row = {
        "instrument_id": instrument_id,
        "instrument_name": instrument_id,
        "instrument_type": "macro",
        "prices_dataset_id": "DAILY_MACRO_INSTRUMENTS_XLSX",
        "open_col": None,
        "high_col": None,
        "low_col": None,
        "close_col": None,
        "volume_col": None,
        "units": UNITS.get(instrument_id),
        "return_model": None,
        "tick_size": None,
        "multiplier": None,
        "currency": None,
        "calendar": None,
        "default_execution_time": None,
        "default_execution_price_model": None,
        "notes": "Macro daily series from Macro_Instruments.xlsx; series_id strings stored in *_col fields.",
    }

    if last_only:
        row["close_col"] = f"{prefix}|last"
    else:
        row["open_col"] = f"{prefix}|open"
        row["high_col"] = f"{prefix}|high"
        row["low_col"] = f"{prefix}|low"
        row["close_col"] = f"{prefix}|close"
        if has_volume:
            row["volume_col"] = f"{prefix}|volume"
    return row


def _build_es_row() -> dict[str, object]:
    return {
        "instrument_id": "ES",
        "instrument_name": "ES",
        "instrument_type": "futures",
        "prices_dataset_id": "DB_ES_TRADES",
        "open_col": None,
        "high_col": None,
        "low_col": None,
        "close_col": None,
        "volume_col": None,
        "units": None,
        "return_model": None,
        "tick_size": 0.25,
        "multiplier": 50,
        "currency": "USD",
        "calendar": "CME",
        "default_execution_time": None,
        "default_execution_price_model": None,
        "notes": (
            "Placeholder. Trades ingested via DB_ES_TRADES into canonical parquet. "
            "Underlying symbols include multiple expiries and spreads. Roll will be configured later "
            "(front quarterly roll; exclude spreads unless explicitly enabled)."
        ),
    }


def _update_or_append(
    ws,
    header: list[str],
    row_dict: dict[str, object],
    existing_rows: dict[str, int],
) -> tuple[str, int]:
    instrument_id = row_dict["instrument_id"]
    if instrument_id in existing_rows:
        row_idx = existing_rows[instrument_id]
        action = "updated"
    else:
        row_idx = ws.max_row + 1
        action = "appended"

    for col_idx, col_name in enumerate(header, start=1):
        if col_name in row_dict:
            ws.cell(row=row_idx, column=col_idx, value=row_dict[col_name])
    return action, row_idx


def main() -> int:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(CONFIG_PATH)

    macro_headers = _load_macro_headers(MACRO_PATH)
    missing_sheets = [s for s in SHEET_TO_INSTRUMENT if s not in macro_headers]
    if missing_sheets:
        raise ValueError(f"Missing sheets in macro workbook: {missing_sheets}")

    wb = load_workbook(CONFIG_PATH)
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet.")
    ws = wb["INSTRUMENTS"]
    header_row = [c.value for c in ws[1]]
    header = [h for h in header_row if h is not None and str(h).strip() != ""]
    if not header:
        raise ValueError("INSTRUMENTS header row is empty.")
    if "instrument_id" not in header:
        raise ValueError("INSTRUMENTS sheet missing instrument_id column.")

    id_col_idx = header.index("instrument_id") + 1
    existing_rows: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=id_col_idx).value
        if cell_value is None:
            continue
        existing_rows[str(cell_value).strip()] = row_idx

    updated = []
    appended = []

    for sheet_name, instrument_id in SHEET_TO_INSTRUMENT.items():
        cols = macro_headers[sheet_name]
        if not cols:
            raise ValueError(f"Empty header row in macro sheet: {sheet_name}")
        data_cols = cols[1:]
        if not data_cols:
            raise ValueError(f"No data columns found in macro sheet: {sheet_name}")
        prefix = data_cols[0]
        last_only = sheet_name in {"SPX_AD", "SPX_PC"} or len(data_cols) == 1
        has_volume = len(data_cols) == 5

        row = _build_macro_row(
            instrument_id=instrument_id,
            prefix=prefix,
            has_volume=has_volume,
            last_only=last_only,
        )
        action, _ = _update_or_append(ws, header, row, existing_rows)
        (updated if action == "updated" else appended).append(instrument_id)

    es_row = _build_es_row()
    action, _ = _update_or_append(ws, header, es_row, existing_rows)
    (updated if action == "updated" else appended).append("ES")

    wb.save(CONFIG_PATH)
    wb.close()

    print(f"Rows updated: {len(updated)}")
    print(f"Rows appended: {len(appended)}")
    print("Updated instrument_ids:", updated)
    print("Appended instrument_ids:", appended)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
