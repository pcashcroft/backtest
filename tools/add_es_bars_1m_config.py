"""
INSTRUCTION HEADER

What this script does (plain English):
- One-time setup: adds the ES_BARS_1M dataset row to the DATASETS sheet in run_config.xlsx.
- Updates the INSTRUMENTS ES row to reference ES_BARS_1M as its OHLCV price source.
- Re-exports the config snapshot so tools/build_derived_bars_1m.py can read the new config.

This is instrument-aware but ES-specific in the row values. To add a second instrument
(e.g. NQ), add another row using the same pattern with a different instrument_id in notes.

Where to run:
- Run from repo root: C:\\Users\\pcash\\OneDrive\\Backtest

How to run:
- C:\\Users\\pcash\\anaconda3\\envs\\backtest\\python.exe tools\\add_es_bars_1m_config.py

What success looks like:
- Prints "Added ES_BARS_1M to DATASETS" (or "already exists, skipping").
- Prints "Updated INSTRUMENTS ES row".
- Prints "Config snapshot re-exported".

Common failures + fixes:
- Permission error on xlsx: close Excel if the workbook is open, then retry.
- "Missing sheet": run tools/make_run_config_xlsx.py first.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("config/run_config.xlsx")

# New DATASETS row for the ES 1-minute derived bars.
# source_path_or_id = source dataset_id (DB_ES_OHLCV_1S).
# instrument_id in notes is parsed by build_derived_bars_1m.py.
_ES_BARS_1M_ROW: dict[str, object] = {
    "dataset_id": "ES_BARS_1M",
    "dataset_type": "derived_bars",
    "source_type": "derived",
    "source_path_or_id": "DB_ES_OHLCV_1S",
    "update_mode": "incremental",
    "date_col": "bar_time",
    "timestamp_tz": "UTC",
    "bar_frequency": "1m",
    "known_time_rule": "event_time",
    "default_availability_lag_days": 0,
    "canonical_table_name": "bars_1m",
    "canonical_partition_keys": "instrument_id,session,date",
    "notes": "instrument_id: ES",
}

# Fields to set on the INSTRUMENTS ES row.
_ES_INSTRUMENT_UPDATES: dict[str, object] = {
    "prices_dataset_id": "ES_BARS_1M",
    "open_col": "open",
    "high_col": "high",
    "low_col": "low",
    "close_col": "close",
    "volume_col": "volume",
}


def _add_datasets_row(wb) -> bool:
    """Add ES_BARS_1M row to DATASETS sheet. Returns True if added, False if already present."""
    if "DATASETS" not in wb.sheetnames:
        raise ValueError("Missing DATASETS sheet in workbook.")

    ws = wb["DATASETS"]
    headers = [c.value for c in ws[1]]

    if "dataset_id" not in headers:
        raise ValueError("DATASETS sheet has no 'dataset_id' header row.")

    id_col = headers.index("dataset_id") + 1
    existing = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 2)}

    if "ES_BARS_1M" in existing:
        return False

    new_row = [_ES_BARS_1M_ROW.get(h) for h in headers]
    ws.append(new_row)
    return True


def _update_instruments_es(wb) -> bool:
    """Update INSTRUMENTS ES row with OHLCV col references. Returns True if updated."""
    if "INSTRUMENTS" not in wb.sheetnames:
        raise ValueError("Missing INSTRUMENTS sheet in workbook.")

    ws = wb["INSTRUMENTS"]
    headers = [c.value for c in ws[1]]

    if "instrument_id" not in headers:
        raise ValueError("INSTRUMENTS sheet has no 'instrument_id' header.")

    id_col = headers.index("instrument_id") + 1

    for row_idx in range(2, ws.max_row + 2):
        if ws.cell(row=row_idx, column=id_col).value == "ES":
            for field, value in _ES_INSTRUMENT_UPDATES.items():
                if field in headers:
                    ws.cell(row=row_idx, column=headers.index(field) + 1, value=value)
            return True

    return False


def main() -> int:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {XLSX_PATH}. Run tools/make_run_config_xlsx.py first."
        )

    wb = load_workbook(XLSX_PATH)

    added = _add_datasets_row(wb)
    print("Added ES_BARS_1M to DATASETS." if added else "ES_BARS_1M already in DATASETS, skipping.")

    updated = _update_instruments_es(wb)
    print("Updated INSTRUMENTS ES row." if updated else "WARNING: ES row not found in INSTRUMENTS.")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Workbook saved: {XLSX_PATH}")

    subprocess.run(
        [sys.executable, "tools/export_config_snapshot.py"],
        check=True,
    )
    print("Config snapshot re-exported.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
