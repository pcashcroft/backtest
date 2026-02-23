from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .schema import HEADERS


def _is_blank_row(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def read_sheets_as_records(xlsx_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path)
    data: dict[str, list[dict[str, Any]]] = {}

    for sheet_name, headers in HEADERS.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if _is_blank_row(list(row)):
                continue
            record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            rows.append(record)
        data[sheet_name] = rows
    return data
